from docxtpl import DocxTemplate
from openpyxl import Workbook
from openpyxl import load_workbook
from docx import Document
from db_requests import db

def edi_export(doctype='sell', date='', number="",
               reciever='', sender='', contract_num='', driver='', proxy='',
               goods=None, reason='',
               pass_num=0, pass_date='', pass_exp_date=''):
    """
    Export to docx and xlsx file. In square brackets are doctypes thar require given param

    :param doctype: type of document - one of [sell(S),pass(P),waybill(W),move(M), write-off(O)]
    :param date: date of document creation [S,M,O,W]
    :param number: number of document [S,M,O,W]
    :param reciever: one who recieves goods [S,M,W]
    :param sender: one who sends goods [S,M,O,W]
    :param goods: dict of goods transfered or dict of a good properties[S,W,O,M,P]
    :param pass_num: passport number [P]
    :param pass_date: date passport issued [P]
    :param pass_exp_date: date passport expires [P]
    :param contract_num: number of contract between organiztions [M,W]
    :param driver: driver name [M,W]
    :param proxy: proxy number [W]
    :param reason: reason for write-off or anything else [O]
    :return: tuple(len=2) names of the files created
    """

    # sell
    if doctype == 'sell':
        wb = load_workbook(rf'templates\attachment_template.xlsx')
        ws = wb.active
        doc = DocxTemplate(r'templates\sell_template.docx')
        goods_list = ''
        total = 0
        ws['B2'] = 'продаже №' + str(number)
        for key, value in goods.items():
            a = [key]
            a.extend(list(value.values()))
            ws.append(a)

            total += value['amount'] * value['price']
            goods_list += f"{key} x {value['amount']} "
        context = {'number': number,
                   'date': date,
                   'customer': reciever,
                   'sum': str(total)+' руб',
                   'seller': sender,
                   'goods': goods_list}
        doc.render(context)
        doc.save(r'EDI\sell_'+number+'_' + date +'.docx')
        wb.save(r'EDI\sell_'+number+'_' + date + '.xlsx')
        return r'EDI\sell_'+number+'_' + date +'.docx', r'EDI\sell_' + date + '.xlsx'

    # passport
    elif doctype == 'pass':
        doc = DocxTemplate(r'templates\pass_template.docx')
        prop_list = '\n'
        for key, value in goods[list(goods.keys())[0]].items():
            if key != 'exp_date':
                prop_list += f"{key} : {str(value)} \n"
        context = {'name': list(goods.keys())[0],
                   'exp_date': goods[list(goods.keys())[0]]['exp_date'],
                   'pass_num': pass_num,
                   'pass_date': pass_date,
                   'pass_exp_date': pass_exp_date,
                   'properties': prop_list}
        doc.render(context)
        doc.save(r'EDI\pass_'+number+'_' + date + '.docx')
        return r'EDI\pass_'+number+'_' + date + '.docx'

    #акт приемки и акт перемещения
    elif 'move' in doctype:
        direction = doctype.split('_')[1]
        wb = load_workbook(rf'templates\attachment_template.xlsx')
        ws = wb.active
        doc = DocxTemplate(rf'templates\move_{direction}_template.docx')
        if direction == 'in':
            ws['B2'] = f'Акту приемки №' + str(number)
        else:
            ws['B2'] = f'Акту перемещения №' + str(number)
        goods_list = ''
        total = 0
        for key, value in goods.items():
            a = [key]
            a.extend(list(value.values()))
            ws.append(a)
            total += value['amount'] * value['price']
            goods_list += f"{key} x {value['amount']} "
        context = {'number': number,
                   'contract_num': contract_num,
                   'date': date,
                   'reciever': reciever,
                   'driver': driver,
                   'sum': str(total) + ' руб',
                   'sender': sender,
                   'goods': goods_list}
        doc.render(context)
        doc.save(rf'EDI\move_{direction}_'+number+'_' + date + '.docx')
        wb.save(rf'EDI\move_{direction}_'+number+'_' + date + '.xlsx')
        return rf'EDI\move_{direction}_'+number+'_' + date + '.docx', rf'EDI\move_{direction}_' \
               + number + '_' + date + '.xlsx'

    # акт списания
    elif doctype == 'write-off':
        wb = load_workbook(rf'templates\attachment_template.xlsx')
        ws = wb.active
        ws['B2'] = f'Акту списания №' + str(number)
        doc = DocxTemplate(r'templates\writeoff_template.docx')
        goods_list = ''
        total = 0
        for key, value in goods.items():
            a = [key]
            a.extend(list(value.values()))
            ws.append(a)
            total += value['amount'] * value['price']
            goods_list += f"{key} x {value['amount']} "
        context = {'number': number,
                   'date': date,
                   'sum': str(total)+' руб',
                   'sender': sender,
                   'goods': goods_list,
                   'reason': reason}
        doc.render(context)
        doc.save(r'EDI\writeoff_'+number+'_' + date + '.docx')
        wb.save(r'EDI\writeoff_'+number+'_' + date + '.xlsx')
        return r'EDI\writeoff_'+number+'_' + date + '.docx', r'EDI\writeoff_'\
               + number + '_' + date + '.xlsx'

    # накладная
    elif doctype == 'waybill':
        wb = load_workbook(rf'templates\attachment_template.xlsx')
        ws = wb.active
        doc = DocxTemplate(rf'templates\waybill_template.docx')
        ws['B2'] = f'Накладной №' + str(number)
        goods_list = ''
        total = 0
        for key, value in goods.items():
            a = [key]
            a.extend(list(value.values()))
            ws.append(a)
            total += value['amount'] * value['price']
            goods_list += f"{key} x {value['amount']} "
        context = {'number': number,
                   'contract_num': contract_num,
                   'date': date,
                   'reciever': reciever,
                   'driver': driver,
                   'sum': str(total) + ' руб',
                   'sender': sender,
                   'goods': goods_list,
                   'proxy': proxy}
        doc.render(context)
        doc.save(r'EDI\waybill_' + number + '_' + date + '.docx')
        wb.save(r'EDI\waybill_' + number + '_' + date + '.xlsx')
        return r'EDI\waybill_' + number + '_' + date + '.docx', r'EDI\writeoff_' \
               + number + '_' + date + '.xlsx'


def edi_import(path: str):
    if 'docx' in path:
        document = Document(path)
        par = document.paragraphs
        doc_name = par[0].text.split('№')[0].strip()
        if 'Паспорт' in doc_name:
            exp_date = par[2].text.split(':')[1],
            pass_num = par[0].text.split('№')[1].strip(),
            pass_date = par[3].text.split(':')[1],
            pass_exp_date = par[4].text.split(':')[1],
            properties = {}
            for i in par[5].text.split('\n')[1:-1]:
                properties[i.split(':')[0]] = i.split(':')[1]
            return 'pass', pass_num[0], pass_date[0], pass_exp_date[0], exp_date[0], properties

    elif 'xlsx' in path:
        wb = load_workbook(filename=path)
        ws = wb.active
        goods_list = []
        for row in ws.iter_rows(min_row=5, max_col=5, max_row=100):
            a = [i.value for i in row]
            if any(a):
                goods_list.append(a)
        return goods_list
    pass
    # doctype = 'sell',
    # date = '',
    # number = "",
    # reciever = '',
    # sender = '',
    # contract_num = '',
    # driver = '',
    # proxy = '',
    # goods = None,
    # reason = '',
    # pass_num = 0,
    # pass_date = '',
    # pass_exp_date = ''


# edi_export(date='12.12.2023',
#            reciever='Андрей',
#            number='1',
#            sender='Костя',
#            goods={'готовый проект': {'amount': 1,
#                                      'price': 1000,
#                                      'mass': 'very heavy'
#                                      }})
#
edi_export(doctype='pass',
           pass_num=1707,
           date='12.13.2023',
           pass_date='11.13.2023',
           pass_exp_date='не ограничен',
           goods={'готовый проект': {'душность': 100,
                                     'стабильность': 0.1,
                                     'масса': 'very heavy',
                                     'exp_date': 'не ограничен'
                                     }})
#
#
# edi_export(number='1',
#            doctype='move_in',
#            contract_num='12312',
#            driver='Ярик',
#            date='12.12.2023',
#            reciever='ООО АЮК',
#            sender='Бабушка',
#            goods={'техзадание': {'amount': 1,
#                                      'price': 1000,
#                                      'mass': 'невыносимая'
#                                  }})
#
# edi_export(number='1',
#            doctype='move_out',
#            contract_num='12312',
#            driver='Ярик',
#            date='12.12.2023',
#            reciever='Бабушка',
#            sender='ООО АЮК',
#            goods={'готовый проект': {'price': 1000,
#                                      'vendor_code': 'R-123',
#                                      'mass': '28 тысяч коммитов! ты дейстовал наверняка',
#                                      'amount': 1
#                                      }
#                   }
#            )
#
# edi_export(number='1',
#            doctype='write-off',
#            date='12.12.2023',
#            sender='ООО АЮК',
#            reason='перегорели',
#            goods={'человеческие мозги': {'amount': 3,
#                                      'price': 1000,
#                                      'mass': '1 кг'
#                                  }})
#
# edi_export(number='1',
#            doctype='write-off',
#            date='12.12.2023',
#            sender='ООО АЮК',
#            reason='перегорели',
#            goods={'человеческие мозги': {'amount': 3,
#                                      'price': 1000,
#                                      'mass': '1 кг'
#                                  }})
#
#
# edi_export(number='1',
#            doctype='waybill',
#            contract_num='12312',
#            driver='Ярик',
#            date='12.12.2023',
#            reciever='Бабушка',
#            sender='ООО АЮК',
#            goods={'готовый проект': {'amount': 1,
#                                      'price': 1000,
#                                      'mass': '28 тысяч коммитов! ты дейстовал наверняка'
#                                      }})


# project = {"вес":10000,
#            "душность": 100,
#            "реализумеость":0,
#            'exp_date':''}
# edi_export(doctype='pass',
#            number='1',
#            date='14.12.2023',
#            pass_num=12313,
#            pass_date='14.10.23',
#            pass_exp_date='10.10.2043',
#            goods={'project': project})

# good_list = edi_import('EDI/move_out_1_12.12.2023.xlsx')
# print(good_list)
# for good in good_list:
#     db.insert('Goods', (good[0], good[1], '', good[2], good[3], '', '', 0, '', good[4], '', ''))
# print(db.get_table_by_name('Goods'))

# pass_list = edi_import(r'EDI/pass__12.13.2023.docx')
# print(pass_list)
# db.insert_docs(doc_type=pass_list[0],  # можно в словарь обернуть тут и в бэке позже
#                prefix='ПА',
#                number=pass_list[1],
#                date='26.12.2023',
#                pass_issued=pass_list[2],
#                pass_expired=pass_list[3],
#                )
# print(db.get_table_by_name('Docs', with_id=True))
